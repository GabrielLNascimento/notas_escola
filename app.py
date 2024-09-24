from pathlib import Path
import os

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROUTE_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROUTE_FOLDER / 'notas.xlsx'

# carregar planilha
workbook: Workbook = load_workbook(WORKBOOK_PATH)
sheet_name = 'Página1'
worksheet: Worksheet = workbook[sheet_name]

coluna_p1 = "B"
coluna_p2 = "C"
coluna_t = "D"

def registrar():
    for i, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        os.system('cls')
        print(f'Matéria.: {row[0]}')
        print(f'p1......: {row[1]}')
        print(f'p2......: {row[2]}')
        print(f'trabalho: {row[3]}')

        p1 = input('Digite a nota da p1: ')
        
        # trocar de matéria rapido
        if p1 == "@":
            continue
        
        p2 = input('Digite a nota da p2: ')
        t = input('Digite a nota do trabalho: ')

        

        if p1 != "!":
            worksheet[f'{coluna_p1}{i}'] = p1
        
        if p2 != "!":
            worksheet[f'{coluna_p2}{i}'] = p2
        
        if t != "!":
            worksheet[f'{coluna_t}{i}'] = t
        
        workbook.save(WORKBOOK_PATH)

def consultar():
    mater = input('digite o nome da matéria: ')

    for row in worksheet.iter_rows(values_only=True, min_row=2):
        if row[0] == mater:
            print(f'Matéria.: {row[0]}')
            print(f'p1......: {row[1]}')
            print(f'p2......: {row[2]}')
            print(f'trabalho: {row[3]}')
            os.system('pause')
            menu()

def listar():
    for row in worksheet.iter_rows(values_only=True, min_row=2):
        print(f'Matéria.: {row[0]}')
        print(f'p1......: {row[1]}')
        print(f'p2......: {row[2]}')
        print(f'trabalho: {row[3]}')
        print('\n')

def calcular_media():
    for i, row in enumerate(worksheet.iter_rows(values_only=True, min_row=2), start=2):
        p1 = worksheet[f'{coluna_p1}{i}'].value
        p2 = worksheet[f'{coluna_p2}{i}'].value
        t = worksheet[f'{coluna_t}{i}'].value
        
        if p1 is None or p2 is None or t is None:
            continue

        if p1 == '' or p2 == '' or t == '':
            continue

        try:
            p1_f = float(p1)
            p2_f = float(p2)
            t_f = float(t)
        except:
            continue
        media = ((p1_f * 2) + (p2_f * 2) + t_f) / 5

        worksheet[f'E{i}'] = media

    
    for row in worksheet.iter_rows(values_only=True, min_row=2):
        print(f'Matéria.: {row[0]}')
        print(f'Média...: {row[4]}')

        print('\n')
    os.system('pause')
    menu()
 


def menu():
    os.system('cls')
    print('1 - registrar nota')
    print('2 - consultar matéria')
    print('3 - listar matérias e notas')
    print('4 - Mostrar médias')

    op = input('Escolha uma opção: ')

    if op not in '1234':
        menu()
    
    match op:
        case '1':
            registrar()
        case '2':
            consultar()
        case '3':
            listar()
        case '4':
            calcular_media()

if __name__ == '__main__':
    menu()